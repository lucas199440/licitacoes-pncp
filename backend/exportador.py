"""
backend/exportador.py
Exportação para Excel com formatação profissional.
"""

import pandas as pd
import os
import io
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXPORTS_DIR = "/tmp/exports"
os.makedirs(EXPORTS_DIR, exist_ok=True)


def exportar_excel(licitacoes: list) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho = os.path.join(EXPORTS_DIR, f"licitacoes_pncp_{ts}.xlsx")

    rows = []
    for i, l in enumerate(licitacoes, 1):
        rows.append({
            "#": i,
            "Objeto": l.get("objeto", ""),
            "Órgão": l.get("orgao", ""),
            "UF": l.get("uf", ""),
            "Município": l.get("municipio", ""),
            "Modalidade": l.get("modalidade", ""),
            "Situação": l.get("situacao", ""),
            "Data Publicação": l.get("data_publicacao", ""),
            "Data Abertura": l.get("data_abertura", ""),
            "Valor Estimado (R$)": float(l.get("valor_estimado") or 0),
            "Link Edital": l.get("link_edital", ""),
            "Favorito": "⭐" if l.get("favoritado") else "",
        })

    df = pd.DataFrame(rows if rows else [{"#": "", "Objeto": "Sem resultados"}])

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Licitações")
        ws = writer.sheets["Licitações"]

        # Cabeçalho
        hfill = PatternFill(start_color="0F2240", end_color="0F2240", fill_type="solid")
        hfont = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
        for ci, _ in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=ci)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Dados
        fpar  = PatternFill(start_color="EBF1F8", end_color="EBF1F8", fill_type="solid")
        fimpar= PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        bord  = Border(bottom=Side(style="thin", color="D0D7E3"))

        col_val = col_link = None
        for ci, cn in enumerate(df.columns, 1):
            if cn == "Valor Estimado (R$)": col_val = ci
            if cn == "Link Edital": col_link = ci

        for ri in range(2, len(df) + 2):
            for ci in range(1, len(df.columns) + 1):
                c = ws.cell(row=ri, column=ci)
                c.fill = fpar if ri % 2 == 0 else fimpar
                c.font = Font(size=9, name="Calibri")
                c.border = bord
                c.alignment = Alignment(vertical="top", wrap_text=True)
                if ci == col_val:
                    c.number_format = 'R$ #,##0.00'
                    c.alignment = Alignment(horizontal="right", vertical="top")
                if ci == col_link and c.value:
                    c.hyperlink = c.value
                    c.font = Font(color="0563C1", underline="single", size=9, name="Calibri")
                    c.value = "Abrir Edital"

        # Larguras
        larguras = {"#":5,"Objeto":55,"Órgão":35,"UF":6,"Município":18,
                    "Modalidade":24,"Situação":20,"Data Publicação":14,
                    "Data Abertura":14,"Valor Estimado (R$)":18,"Link Edital":14,"Favorito":8}
        for ci, cn in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = larguras.get(cn, 15)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # Aba Resumo
        wr = writer.book.create_sheet("Resumo")
        wr.sheet_view.showGridLines = False
        wr["A1"] = "📊 RESUMO — LICITAÇÕES PNCP"
        wr["A1"].font = Font(bold=True, size=13, color="0F2240", name="Calibri")
        wr["A3"] = "Gerado em:"
        wr["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        wr["A4"] = "Total de Licitações:"
        wr["B4"] = len(licitacoes)
        wr["A5"] = "Valor Total Estimado:"
        wr["B5"] = sum(float(l.get("valor_estimado") or 0) for l in licitacoes)
        wr["B5"].number_format = 'R$ #,##0.00'

        mods = {}
        for l in licitacoes:
            m = l.get("modalidade") or "Não informado"
            mods[m] = mods.get(m, 0) + 1
        wr["A7"] = "Por Modalidade:"
        wr["A7"].font = Font(bold=True, name="Calibri")
        for i, (m, q) in enumerate(sorted(mods.items(), key=lambda x:-x[1]), 8):
            wr[f"A{i}"] = m; wr[f"B{i}"] = q

        wr.column_dimensions["A"].width = 35
        wr.column_dimensions["B"].width = 20

    return caminho
